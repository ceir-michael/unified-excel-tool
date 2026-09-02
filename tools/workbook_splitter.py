from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook

from core.excel_helpers import copy_cell
from core.file_utils import safe_file_name, safe_sheet_name


def run_workbook_splitter(params: dict, log: Callable[[str], None]) -> str:
    """Create one workbook per unique split-column value."""
    input_path = Path(params["input_file"])
    output_folder = Path(params["output_folder"])
    header_row = params["header_row"]
    split_column = params["split_column"]
    tab_column = params.get("tab_column") or None

    output_folder.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_path)
    source = workbook.active

    headers: dict[str, int] = {}
    for column_number in range(1, source.max_column + 1):
        value = source.cell(row=header_row, column=column_number).value
        if value is not None:
            headers[str(value).strip()] = column_number

    if split_column not in headers:
        raise ValueError(f"Split column not found: {split_column}")
    if tab_column and tab_column not in headers:
        raise ValueError(f"Tab column not found: {tab_column}")
    if tab_column == split_column:
        raise ValueError("The split column and tab column must be different.")

    split_index = headers[split_column]
    tab_index = headers[tab_column] if tab_column else None

    grouped: dict = {}
    for row_number in range(header_row + 1, source.max_row + 1):
        split_value = source.cell(row=row_number, column=split_index).value
        tab_value = (
            source.cell(row=row_number, column=tab_index).value
            if tab_index
            else None
        )
        grouped.setdefault(split_value, {}).setdefault(tab_value, []).append(row_number)

    created_files: list[Path] = []
    used_file_names: set[str] = set()

    for split_value, tab_groups in grouped.items():
        output_workbook = Workbook()
        output_workbook.remove(output_workbook.active)
        used_sheet_names: set[str] = set()

        for tab_value, source_rows in tab_groups.items():
            sheet_value = tab_value if tab_column else split_value
            sheet_name = safe_sheet_name(sheet_value, used_sheet_names)
            log(
                f"{split_value}: writing tab '{sheet_name}' "
                f"({len(source_rows)} row(s))"
            )

            target = output_workbook.create_sheet(sheet_name)
            target.freeze_panes = "A2"
            target.sheet_view.showGridLines = source.sheet_view.showGridLines

            for column_number in range(1, source.max_column + 1):
                source_header = source.cell(row=header_row, column=column_number)
                letter = source_header.column_letter
                target.column_dimensions[letter].width = (
                    source.column_dimensions[letter].width
                )
                copy_cell(source_header, target.cell(row=1, column=column_number))

            for target_row, source_row in enumerate(source_rows, start=2):
                for column_number in range(1, source.max_column + 1):
                    copy_cell(
                        source.cell(row=source_row, column=column_number),
                        target.cell(row=target_row, column=column_number),
                    )

            target.auto_filter.ref = target.dimensions

        base_name = f"{input_path.stem} - {safe_file_name(split_value)}"
        file_name = f"{base_name}.xlsx"
        counter = 1

        while (
            file_name.casefold() in used_file_names
            or (output_folder / file_name).exists()
        ):
            file_name = f"{base_name}_{counter}.xlsx"
            counter += 1

        used_file_names.add(file_name.casefold())
        output_path = output_folder / file_name
        output_workbook.save(output_path)
        created_files.append(output_path)
        log(f"Saved {output_path.name}")

    return f"Created {len(created_files)} workbook(s) in {output_folder}"
